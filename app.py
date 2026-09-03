#!/usr/bin/env python3
"""Webapp för att jämföra Excel- och CSV-filer och hitta gemensamma rader."""

import collections
import csv
import datetime
import io
import os
import re
import openpyxl
from flask import Flask, render_template_string, request, Response, Blueprint, redirect

_DATETIME_RE = re.compile(r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}")

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

bp = Blueprint("crossmatch", __name__)

HTML = """
<!DOCTYPE html>
<html lang="sv">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>CrossMatch</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { font-family: system-ui, sans-serif; background: #f5f5f5; color: #333; padding: 2rem; }
        h1 { margin-bottom: 1.5rem; }
        .drop-zone {
            border: 3px dashed #aaa; border-radius: 12px; padding: 3rem;
            text-align: center; background: #fff; cursor: pointer;
            transition: border-color .2s, background .2s; margin-bottom: 1rem;
        }
        .drop-zone.over { border-color: #2563eb; background: #eff6ff; }
        .drop-zone p { font-size: 1.1rem; color: #666; }
        .file-list { margin: 1rem 0; }
        .file-list span {
            display: inline-block; background: #e0e7ff; color: #3730a3;
            padding: .3rem .7rem; border-radius: 6px; margin: .2rem; font-size: .9rem;
        }
        .file-list span button {
            background: none; border: none; color: #6366f1; cursor: pointer;
            font-weight: bold; margin-left: .3rem;
        }
        .loading { color: #2563eb; font-style: italic; margin: .5rem 0; display: none; }
        .results { margin-top: 2rem; }
        .results h2 { margin-bottom: .5rem; }
        .summary { background: #fff; padding: 1rem 1.5rem; border-radius: 8px; margin-bottom: 1rem; }
        .summary p { margin: .3rem 0; }
        table {
            width: 100%; border-collapse: collapse; background: #fff;
            border-radius: 8px; overflow: hidden; font-size: .85rem;
        }
        th, td { padding: .5rem .7rem; border-bottom: 1px solid #eee; text-align: left; }
        th { background: #f8fafc; position: sticky; top: 0; }
        .table-wrap { max-height: 500px; overflow: auto; border-radius: 8px; border: 1px solid #ddd; }
        .download-btn {
            display: inline-block; margin-top: 1rem; padding: .5rem 1.2rem;
            background: #16a34a; color: #fff; border-radius: 6px;
            text-decoration: none; font-size: .95rem;
        }
        .download-btn:hover { background: #15803d; }
        .download-btn.excel { background: #1d6a96; margin-left: .5rem; }
        .download-btn.excel:hover { background: #155a80; }
        .pair { display: inline-block; background: #fef3c7; padding: .2rem .6rem; border-radius: 4px; margin: .15rem; font-size: .85rem; }
        .col-tag { display: inline-block; background: #dcfce7; color: #166534; padding: .15rem .5rem; border-radius: 4px; margin: .1rem; font-size: .8rem; }
    </style>
</head>
<body>
    <h1>CrossMatch</h1>

    <form id="form" enctype="multipart/form-data">
        <div class="drop-zone" id="dropZone">
            <p>Dra och släpp Excel- eller CSV-filer här, eller klicka för att välja</p>
            <input type="file" name="files" id="fileInput" multiple accept=".xlsx,.xlsm,.xls,.csv" style="display:none">
        </div>
        <div class="file-list" id="fileList"></div>
        <div class="loading" id="loading">Jämför filer...</div>
    </form>

    <div id="results">
    {% if results %}
    <div class="results">
        <div class="summary">
            <h2>Resultat</h2>
            <p><strong>Jämförda kolumner:</strong>
                {% for col in results.shared_cols %}<span class="col-tag">{{ col }}</span>{% endfor %}
            </p>
            <p style="margin-top:.5rem"><strong>Antal filer:</strong> {{ results.file_count }}</p>
            {% for fname, count in results.file_stats %}
            <p>&nbsp;&nbsp;{{ fname }}: {{ count }} unika rader</p>
            {% endfor %}
            <p style="margin-top:.5rem"><strong>Gemensamma rader (finns i minst 2 filer):</strong> {{ results.common_count }}</p>
            {% if results.pairwise %}
            <p style="margin-top:.5rem"><strong>Parvis överlapp:</strong></p>
            {% for pair, count in results.pairwise %}
            <span class="pair">{{ pair }}: {{ count }}</span>
            {% endfor %}
            {% endif %}
        </div>

        {% if results.common_count > 0 %}
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Källa</th>
                        {% for col in results.headers %}
                        <th>{{ col }}</th>
                        {% endfor %}
                    </tr>
                </thead>
                <tbody>
                    {% for row in results.rows %}
                    <tr>
                        {% for cell in row %}
                        <td>{{ cell }}</td>
                        {% endfor %}
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        <a class="download-btn" href="{{ url_for('crossmatch.download') }}" target="_blank">Ladda ner som CSV</a>
        <a class="download-btn excel" href="{{ url_for('crossmatch.download_excel') }}" target="_blank">Ladda ner som Excel</a>
        {% endif %}
    </div>
    {% endif %}
    </div>

    <script>
        const dropZone = document.getElementById('dropZone');
        const fileInput = document.getElementById('fileInput');
        const fileList = document.getElementById('fileList');
        const loading = document.getElementById('loading');
        let storedFiles = [];

        dropZone.addEventListener('click', () => fileInput.click());
        dropZone.addEventListener('dragover', e => { e.preventDefault(); dropZone.classList.add('over'); });
        dropZone.addEventListener('dragleave', () => dropZone.classList.remove('over'));
        dropZone.addEventListener('drop', e => {
            e.preventDefault();
            dropZone.classList.remove('over');
            addFiles(e.dataTransfer.files);
        });
        fileInput.addEventListener('change', () => addFiles(fileInput.files));

        function addFiles(newFiles) {
            for (const f of newFiles) {
                if (!storedFiles.some(s => s.name === f.name && s.size === f.size)) {
                    storedFiles.push(f);
                }
            }
            renderList();
            if (storedFiles.length >= 2) compare();
        }

        function removeFile(idx) {
            storedFiles.splice(idx, 1);
            renderList();
            if (storedFiles.length >= 2) compare();
            else document.getElementById('results').innerHTML = '';
        }

        function renderList() {
            fileList.innerHTML = storedFiles.map((f, i) =>
                `<span>${f.name} <button type="button" onclick="removeFile(${i})">&#x2715;</button></span>`
            ).join('');
        }

        function compare() {
            loading.style.display = 'block';
            const fd = new FormData();
            storedFiles.forEach(f => fd.append('files', f));

            fetch('{{ url_for("crossmatch.index") }}', { method: 'POST', body: fd })
                .then(r => r.text())
                .then(html => {
                    loading.style.display = 'none';
                    const doc = new DOMParser().parseFromString(html, 'text/html');
                    document.getElementById('results').innerHTML =
                        doc.getElementById('results').innerHTML;
                });
        }
    </script>
</body>
</html>
"""

# Global store for the last comparison result (for download)
last_result = {}


def parse_file(file_storage):
    """Return (rows, headers) where rows are dicts with raw Python values (preserving types)."""
    filename = file_storage.filename or ""
    ext = os.path.splitext(filename)[1].lower()

    if ext in (".xlsx", ".xlsm", ".xls"):
        wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(c) if c is not None else "" for c in next(rows_iter, [])]
        rows = [
            {headers[i]: v for i, v in enumerate(row)}
            for row in rows_iter
            if any(v is not None for v in row)
        ]
        wb.close()
        return rows, headers

    # CSV fallback — all values are already strings
    raw = file_storage.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    first_line = text.split("\n", 1)[0]
    delimiter = ";" if ";" in first_line else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows = [r for r in reader if any(v.strip() for v in r.values())]
    return rows, list(reader.fieldnames) if reader.fieldnames else []


def cell_str(v):
    """Convert a cell value to a comparable string."""
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.isoformat()
    return str(v).strip()


def is_datetime_col(col, all_data):
    """Return True if the column contains datetime values in any file.
    Handles both Excel datetime objects and CSV datetime strings."""
    for rows, _ in all_data.values():
        for row in rows:
            v = row.get(col)
            if isinstance(v, datetime.datetime):
                return True
            if isinstance(v, str) and _DATETIME_RE.match(v):
                return True
    return False


def has_cross_file_overlap(col, all_data):
    """Return True if the column has at least one common non-empty value across ALL files.
    Kept strict so that per-row unique IDs (like Subscriber ID) aren't used as key columns."""
    value_sets = [
        {cell_str(row.get(col)) for row in rows} - {""}
        for rows, _ in all_data.values()
    ]
    return bool(set.intersection(*value_sets)) if value_sets else False


@bp.route("/", methods=["GET", "POST"])
def index():
    results = None

    if request.method == "POST":
        files = request.files.getlist("files")
        files = [f for f in files if f.filename]

        if len(files) >= 2:
            all_data = {}
            all_header_sets = []
            first_headers = None
            for f in files:
                rows, hdrs = parse_file(f)
                all_data[f.filename] = (rows, hdrs)
                all_header_sets.append(set(hdrs))
                if first_headers is None:
                    first_headers = hdrs

            # Key columns: shared across all files, non-datetime,
            # and must have at least one common value across files.
            # This excludes per-row unique IDs (like Subscriber ID) that never overlap.
            shared_set = set.intersection(*all_header_sets)
            shared_cols = [
                h for h in first_headers
                if h in shared_set
                and not is_datetime_col(h, all_data)
                and has_cross_file_overlap(h, all_data)
            ]

            def row_key(row):
                return tuple(cell_str(row.get(c)) for c in shared_cols)

            keys_per_file = {}
            for fname, (rows, _) in all_data.items():
                keys_per_file[fname] = {row_key(r) for r in rows}

            # Keys that appear in at least 2 files
            key_file_count = collections.Counter()
            for keys in keys_per_file.values():
                key_file_count.update(keys)
            common_keys = {k for k, n in key_file_count.items() if n >= 2}

            # Pairwise overlap
            filenames = list(keys_per_file.keys())
            pairwise = []
            for i in range(len(filenames)):
                for j in range(i + 1, len(filenames)):
                    overlap = keys_per_file[filenames[i]] & keys_per_file[filenames[j]]
                    pairwise.append((f"{filenames[i]} & {filenames[j]}", len(overlap)))

            # Result rows: all columns from first file, values converted to strings
            result_rows = []
            for fname, (rows, _) in all_data.items():
                for row in rows:
                    if row_key(row) in common_keys:
                        result_rows.append([fname] + [cell_str(row.get(h)) for h in first_headers])

            results = {
                "file_count": len(all_data),
                "file_stats": [(fn, len(keys_per_file[fn])) for fn in filenames],
                "common_count": len(common_keys),
                "pairwise": pairwise,
                "shared_cols": shared_cols,
                "headers": first_headers,
                "rows": result_rows,
            }

            last_result["headers"] = first_headers
            last_result["rows"] = result_rows

    return render_template_string(HTML, results=results)


@bp.route("/download")
def download():
    if not last_result.get("rows"):
        return "Inget resultat att ladda ner", 404

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Källa"] + last_result["headers"])
    for row in last_result["rows"]:
        writer.writerow(row)

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=gemensamma.csv"},
    )


@bp.route("/download-excel")
def download_excel():
    if not last_result.get("rows"):
        return "Inget resultat att ladda ner", 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gemensamma värden"
    ws.append(["Källa"] + last_result["headers"])
    for row in last_result["rows"]:
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=gemensamma.xlsx"},
    )


app.register_blueprint(bp, url_prefix="/crossmatch")


@app.route("/")
def root_redirect():
    return redirect("/crossmatch/")


if __name__ == "__main__":
    app.run(host="0.0.0.0", debug=True, port=5002)
